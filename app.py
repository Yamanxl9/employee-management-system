from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_pymongo import PyMongo
from datetime import datetime, timedelta, timezone
import pandas as pd
import io
import os
import json
import jwt
import re
from functools import wraps
from dotenv import load_dotenv
from nationalities import NATIONALITIES
from bson import ObjectId
from werkzeug.security import generate_password_hash, check_password_hash
from nationalities import get_nationality_name, get_all_nationalities

import logging

# إعداد التسجيل
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# تحميل متغيرات البيئة
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-here')
app.config['MONGO_URI'] = os.getenv('MONGODB_URI', 'mongodb+srv://yamanxl9:SgIhE4u0FbzRhbyp@cluster0.lqokm.mongodb.net/employee_management?retryWrites=true&w=majority&appName=Cluster0')

# إعداد MongoDB مع معالجة الأخطاء
try:
    mongo = PyMongo(app)
    # اختبار الاتصال
    mongo.db.test_connection.find_one()
    logger.info("✅ MongoDB connection successful!")
except Exception as e:
    logger.error(f"❌ MongoDB connection failed: {e}")
    mongo = None

# دالة للتحقق من اتصال MongoDB
def check_mongodb_connection():
    """التحقق من حالة اتصال MongoDB"""
    if not mongo:
        return False, "MongoDB connection not initialized"
    
    try:
        # اختبار بسيط للاتصال
        mongo.db.admin.command('ping')
        return True, "MongoDB connection is healthy"
    except Exception as e:
        return False, f"MongoDB connection error: {str(e)}"

# دالة لتسجيل الأنشطة في Audit Log
def log_activity(action, details, user_id=None):
    """تسجيل نشاط في سجل التدقيق"""
    if not mongo:
        logger.warning("Cannot log activity: MongoDB not connected")
        return
        
    try:
        audit_log = {
            'timestamp': datetime.now(),
            'action': action,
            'details': details,
            'user_id': user_id or session.get('user_id', 'مجهول'),
            'ip_address': request.remote_addr,
            'user_agent': request.headers.get('User-Agent', '')
        }
        mongo.db.audit_logs.insert_one(audit_log)
    except Exception as e:
        logger.error(f"Error logging activity: {e}")

# Helper functions for MongoDB
def serialize_doc(doc):
    """تحويل وثيقة MongoDB إلى قاموس قابل للتسلسل"""
    if doc is None:
        return None
    if isinstance(doc, list):
        return [serialize_doc(item) for item in doc]
    
    if '_id' in doc:
        doc['_id'] = str(doc['_id'])
    
    # تحويل datetime objects
    for key, value in doc.items():
        if isinstance(value, datetime):
            doc[key] = value.isoformat()
        elif isinstance(value, ObjectId):
            doc[key] = str(value)
    
    return doc

def get_employee_status(employee):
    """حساب حالة الجواز والبطاقة للموظف"""
    passport_status = 'available' if employee.get('pass_no') else 'missing'
    passport_text = 'متوفر' if employee.get('pass_no') else 'غير متوفر'
    passport_class = 'success' if employee.get('pass_no') else 'danger'
    
    # حالة البطاقة
    if not employee.get('card_no'):
        card_status, card_text, card_class = 'missing', 'غير متوفرة', 'danger'
    elif not employee.get('card_expiry_date'):
        card_status, card_text, card_class = 'no_expiry', 'بدون تاريخ انتهاء', 'warning'
    else:
        expiry_date = employee['card_expiry_date']
        if isinstance(expiry_date, str):
            expiry_date = datetime.fromisoformat(expiry_date.replace('Z', '+00:00'))
        
        # التأكد من أن expiry_date يحتوي على timezone
        if expiry_date.tzinfo is None:
            expiry_date = expiry_date.replace(tzinfo=timezone.utc)
        
        today = datetime.now(timezone.utc)
        if expiry_date < today:
            card_status, card_text, card_class = 'expired', 'منتهية الصلاحية', 'danger'
        elif expiry_date < today + timedelta(days=90):
            card_status, card_text, card_class = 'expiring_soon', 'تنتهي قريباً', 'warning'
        else:
            card_status, card_text, card_class = 'valid', 'سارية', 'success'
    
    return {
        'passport_status': passport_status,
        'passport_text': passport_text,
        'passport_class': passport_class,
        'card_status': card_status,
        'card_text': card_text,
        'card_class': card_class
    }

# دوال المصادقة والحماية
def generate_token(user_id):
    """توليد JWT token للمستخدم"""
    payload = {
        'user_id': str(user_id),
        'exp': datetime.now(timezone.utc) + timedelta(days=7)  # صالح لمدة أسبوع
    }
    return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')

def verify_token(token):
    """التحقق من صحة JWT token"""
    try:
        payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
        return payload['user_id']
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

def require_auth(f):
    """decorator للتحقق من تسجيل الدخول"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if token and token.startswith('Bearer '):
            token = token[7:]  # إزالة "Bearer "
            user_id = verify_token(token)
            if user_id:
                request.current_user_id = user_id
                return f(*args, **kwargs)
        
        return jsonify({'error': 'تسجيل الدخول مطلوب'}), 401
    return decorated_function

def init_admin_user():
    """إنشاء مستخدم إداري افتراضي"""
    admin = mongo.db.users.find_one({'username': 'admin'})
    if not admin:
        admin_data = {
            'username': 'admin',
            'password': generate_password_hash('admin123'),
            'role': 'admin',
            'created_at': datetime.now()
        }
        mongo.db.users.insert_one(admin_data)
        print("تم إنشاء المستخدم الإداري: admin / admin123")

# API للتحقق من حالة النظام
@app.route('/api/health')
def health_check():
    """فحص حالة النظام والاتصال بقاعدة البيانات"""
    is_connected, message = check_mongodb_connection()
    
    return jsonify({
        'status': 'healthy' if is_connected else 'unhealthy',
        'mongodb': {
            'connected': is_connected,
            'message': message
        },
        'timestamp': datetime.now().isoformat(),
        'version': '2.0'
    }), 200 if is_connected else 503

# الصفحة الرئيسية
@app.route('/')
def index():
    # السماح بالوصول للصفحة الرئيسية، والتحقق من المصادقة سيتم في JavaScript
    return render_template('index.html')

# صفحة سجل التدقيق
@app.route('/audit-logs')
def audit_logs_page():
    # التحقق من وجود token في cookies أو localStorage سيتم في JavaScript
    return render_template('audit_logs.html')

# صفحة تسجيل الدخول
@app.route('/login')
def login_page():
    return render_template('auth/login.html')

# صفحة اختبار المصادقة
@app.route('/test-auth')
def test_auth():
    return render_template('test_auth.html')

# صفحة التشخيص السريع
@app.route('/debug')
def debug_page():
    return render_template('debug.html')

# API تسجيل الدخول
@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'error': 'اسم المستخدم وكلمة المرور مطلوبان'}), 400
        
        # البحث عن المستخدم
        user = mongo.db.users.find_one({'username': username})
        
        if not user or not check_password_hash(user['password'], password):
            return jsonify({'error': 'اسم المستخدم أو كلمة المرور غير صحيحة'}), 401
        
        # توليد JWT token
        token = generate_token(user['_id'])
        
        # تسجيل النشاط
        log_activity('تسجيل دخول', f'تسجيل دخول ناجح للمستخدم: {username}', str(user['_id']))
        
        # إعداد بيانات المستخدم للإرجاع (بدون كلمة المرور)
        user_data = {
            'id': str(user['_id']),
            'username': user['username'],
            'role': user.get('role', 'user')
        }
        
        return jsonify({
            'token': token,
            'user': user_data,
            'message': 'تم تسجيل الدخول بنجاح'
        })
        
    except Exception as e:
        return jsonify({'error': 'خطأ في الخادم'}), 500

# API تسجيل الخروج
@app.route('/api/logout', methods=['POST'])
@require_auth
def logout():
    return jsonify({'message': 'تم تسجيل الخروج بنجاح'})

# API للتحقق من صحة Token
@app.route('/api/verify-token', methods=['POST'])
def verify_token_route():
    try:
        data = request.get_json()
        token = data.get('token')
        
        if not token:
            return jsonify({'valid': False}), 400
            
        user_id = verify_token(token)
        if user_id:
            user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
            if user:
                user_data = {
                    'id': str(user['_id']),
                    'username': user['username'],
                    'role': user.get('role', 'user')
                }
                return jsonify({'valid': True, 'user': user_data})
        
        return jsonify({'valid': False}), 401
        
    except Exception as e:
        return jsonify({'valid': False}), 500

# API للبحث عن الموظفين
@app.route('/api/search')
@require_auth
def search_employees():
    query = request.args.get('query', '').strip()
    nationality = request.args.get('nationality', '')
    company = request.args.get('company', '')
    job = request.args.get('job', '')
    department = request.args.get('department', '')
    passport_status = request.args.get('passport_status', '')
    card_status = request.args.get('card_status', '')
    emirates_id_status = request.args.get('emirates_id_status', '')
    residence_status = request.args.get('residence_status', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    
    # بناء الاستعلام
    filter_query = {}
    
    if query:
        # البحث الشامل والمتطور في جميع الحقول المهمة
        search_conditions = [
            # البحث في أسماء الموظفين
            {'staff_name': {'$regex': query, '$options': 'i'}},
            {'staff_name_ara': {'$regex': query, '$options': 'i'}},
            # البحث في أرقام الموظفين والوثائق
            {'staff_no': {'$regex': query, '$options': 'i'}},
            {'pass_no': {'$regex': query, '$options': 'i'}},
            {'card_no': {'$regex': query, '$options': 'i'}},
            # البحث في الهوية الإماراتية ورقم الإقامة
            {'emirates_id': {'$regex': query, '$options': 'i'}},
            {'residence_no': {'$regex': query, '$options': 'i'}},
            # البحث في الجنسية
            {'nationality_code': {'$regex': query, '$options': 'i'}}
        ]
        
        # البحث في أسماء الشركات
        companies_matching = list(mongo.db.companies.find({
            '$or': [
                {'company_name_eng': {'$regex': query, '$options': 'i'}},
                {'company_name_ara': {'$regex': query, '$options': 'i'}},
                {'company_code': {'$regex': query, '$options': 'i'}}
            ]
        }))
        
        # إضافة رموز الشركات المطابقة للبحث
        if companies_matching:
            company_codes = [comp['company_code'] for comp in companies_matching]
            search_conditions.append({'company_code': {'$in': company_codes}})
        
        # البحث في أسماء الوظائف
        jobs_matching = list(mongo.db.jobs.find({
            '$or': [
                {'job_eng': {'$regex': query, '$options': 'i'}},
                {'job_ara': {'$regex': query, '$options': 'i'}}
            ]
        }))
        
        # إضافة رموز الوظائف المطابقة للبحث
        if jobs_matching:
            job_codes = [job['job_code'] for job in jobs_matching]
            search_conditions.append({'job_code': {'$in': job_codes}})
        
        # البحث في أسماء الجنسيات الكاملة (Turkish, تركي، etc.)
        matching_nationality_codes = []
        
        for code, names in NATIONALITIES.items():
            en_match = query.lower() in names['en'].lower()
            ar_match = query.lower() in names['ar'].lower()
            code_match = query.lower() in code.lower()
            
            if en_match or ar_match or code_match:
                matching_nationality_codes.append(code)
        
        # إضافة رموز الجنسيات المطابقة للبحث
        if matching_nationality_codes:
            search_conditions.append({'nationality_code': {'$in': matching_nationality_codes}})
        
        filter_query['$or'] = search_conditions
    
    if nationality:
        # البحث الذكي في حقل الجنسية - يدعم الأسماء الكاملة والرموز
        matching_nationality_codes = []
        for code, names in NATIONALITIES.items():
            en_match = nationality.lower() in names['en'].lower()
            ar_match = nationality.lower() in names['ar'].lower()
            code_match = nationality.lower() in code.lower()
            
            if en_match or ar_match or code_match:
                matching_nationality_codes.append(code)
        
        if matching_nationality_codes:
            # البحث بالرموز المطابقة
            filter_query['nationality_code'] = {'$in': matching_nationality_codes}
        else:
            # البحث العادي إذا لم توجد مطابقات
            filter_query['nationality_code'] = {'$regex': nationality, '$options': 'i'}
    
    
    if company:
        # تحسين البحث في الشركة لدعم النص الجزئي
        filter_query['company_code'] = {'$regex': company, '$options': 'i'}
    
    if job:
        try:
            filter_query['job_code'] = int(job)
        except (ValueError, TypeError):
            pass  # تجاهل القيم غير الصحيحة
    
    if department:
        filter_query['department_code'] = {'$regex': department, '$options': 'i'}
    
    # إضافة فلاتر حالة الجواز والبطاقة لقاعدة البيانات مباشرة
    if passport_status == 'missing':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({'$or': [{'pass_no': {'$exists': False}}, {'pass_no': None}, {'pass_no': ''}]})
    elif passport_status == 'available':
        filter_query['pass_no'] = {'$exists': True, '$ne': None, '$ne': ''}
    
    if card_status == 'missing':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({'$or': [{'card_no': {'$exists': False}}, {'card_no': None}, {'card_no': ''}]})
    elif card_status == 'expired':
        # البطاقات المنتهية الصلاحية (لها تاريخ انتهاء ومنتهية)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'card_no': {'$exists': True, '$ne': None, '$ne': ''},
            'card_expiry_date': {'$exists': True, '$ne': None, '$lt': datetime.now()}
        })
    elif card_status == 'expiring_soon':
        # البطاقات التي ستنتهي خلال 90 يوم
        future_date = datetime.now() + timedelta(days=90)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'card_no': {'$exists': True, '$ne': None, '$ne': ''},
            'card_expiry_date': {'$exists': True, '$ne': None, '$gte': datetime.now(), '$lt': future_date}
        })
    elif card_status == 'valid':
        # البطاقات السارية (لها تاريخ انتهاء ولا تزال سارية)
        future_date = datetime.now() + timedelta(days=90)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'card_no': {'$exists': True, '$ne': None, '$ne': ''},
            'card_expiry_date': {'$exists': True, '$ne': None, '$gte': future_date}
        })
    elif card_status == 'no_expiry':
        # البطاقات بدون تاريخ انتهاء (لها رقم بطاقة ولكن لا يوجد تاريخ انتهاء)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'card_no': {'$exists': True, '$ne': None, '$ne': ''},
            '$or': [{'card_expiry_date': {'$exists': False}}, {'card_expiry_date': None}]
        })
    
    # حساب pagination
    skip = (page - 1) * per_page
    
    # إضافة فلاتر حالة الهوية الإماراتية
    if emirates_id_status == 'missing':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({'$or': [{'emirates_id': {'$exists': False}}, {'emirates_id': None}, {'emirates_id': ''}]})
    elif emirates_id_status == 'expired':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'emirates_id': {'$exists': True, '$ne': None, '$ne': ''},
            'emirates_id_expiry': {'$exists': True, '$ne': None, '$lt': datetime.now()}
        })
    elif emirates_id_status == 'expiring_soon':
        future_date = datetime.now() + timedelta(days=90)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'emirates_id': {'$exists': True, '$ne': None, '$ne': ''},
            'emirates_id_expiry': {'$exists': True, '$ne': None, '$gte': datetime.now(), '$lt': future_date}
        })
    elif emirates_id_status == 'valid':
        future_date = datetime.now() + timedelta(days=90)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'emirates_id': {'$exists': True, '$ne': None, '$ne': ''},
            'emirates_id_expiry': {'$exists': True, '$ne': None, '$gte': future_date}
        })
    elif emirates_id_status == 'no_expiry':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'emirates_id': {'$exists': True, '$ne': None, '$ne': ''},
            '$or': [{'emirates_id_expiry': {'$exists': False}}, {'emirates_id_expiry': None}]
        })
    
    # إضافة فلاتر حالة الإقامة
    if residence_status == 'missing':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({'$or': [{'residence_no': {'$exists': False}}, {'residence_no': None}, {'residence_no': ''}]})
    elif residence_status == 'expired':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'residence_no': {'$exists': True, '$ne': None, '$ne': ''},
            'residence_expiry_date': {'$exists': True, '$ne': None, '$lt': datetime.now()}
        })
    elif residence_status == 'expiring_soon':
        future_date = datetime.now() + timedelta(days=90)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'residence_no': {'$exists': True, '$ne': None, '$ne': ''},
            'residence_expiry_date': {'$exists': True, '$ne': None, '$gte': datetime.now(), '$lt': future_date}
        })
    elif residence_status == 'valid':
        future_date = datetime.now() + timedelta(days=90)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'residence_no': {'$exists': True, '$ne': None, '$ne': ''},
            'residence_expiry_date': {'$exists': True, '$ne': None, '$gte': future_date}
        })
    elif residence_status == 'no_expiry':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'residence_no': {'$exists': True, '$ne': None, '$ne': ''},
            '$or': [{'residence_expiry_date': {'$exists': False}}, {'residence_expiry_date': None}]
        })
    
    employees = list(mongo.db.employees.find(filter_query).skip(skip).limit(per_page))
    total = mongo.db.employees.count_documents(filter_query)
    
    # إضافة معلومات الشركة والوظيفة والجنسية
    results = []
    for emp in employees:
        # جلب معلومات الشركة
        company_info = mongo.db.companies.find_one({'company_code': emp.get('company_code')})
        # جلب معلومات الوظيفة
        job_info = mongo.db.jobs.find_one({'job_code': emp.get('job_code')})
        
        emp_dict = serialize_doc(emp)
        emp_dict.update(get_employee_status(emp))
        
        if company_info:
            emp_dict['company_eng'] = company_info.get('company_name_eng', '')
            emp_dict['company_ara'] = company_info.get('company_name_ara', '')
        
        if job_info:
            emp_dict['job_eng'] = job_info.get('job_eng', '')
            emp_dict['job_ara'] = job_info.get('job_ara', '')
        
        # إضافة اسم الجنسية للعرض (إزالة التكرار)
        nationality_code = emp.get('nationality_code', '')
        if nationality_code:
            # إذا كان nationality_code هو رمز (مثل TR, IN)
            if nationality_code in NATIONALITIES:
                emp_dict['nationality_display'] = NATIONALITIES[nationality_code]['ar']
            else:
                # إذا كان nationality_code هو اسم كامل بالفعل (مثل "تركي")
                emp_dict['nationality_display'] = nationality_code
        else:
            emp_dict['nationality_display'] = 'غير محدد'
            
        results.append(emp_dict)
    
    pages = (total + per_page - 1) // per_page
    
    return jsonify({
        'employees': results,
        'total': total,
        'pages': pages,
        'current_page': page,
        'has_next': page < pages,
        'has_prev': page > 1
    })

# API لإضافة موظف جديد
@app.route('/api/employees', methods=['POST'])
@require_auth
def add_employee():
    try:
        data = request.get_json()
        
        # التحقق من البيانات المطلوبة
        required_fields = ['staff_no', 'staff_name', 'staff_name_ara', 'job_code', 'nationality_code', 'company_code']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'حقل {field} مطلوب'}), 400
        
        # التحقق من عدم وجود موظف بنفس الرقم
        existing = mongo.db.employees.find_one({'staff_no': data['staff_no']})
        if existing:
            return jsonify({'error': 'رقم الموظف موجود مسبقاً'}), 400
        
        # تحويل تاريخ انتهاء البطاقة
        if data.get('card_expiry_date'):
            try:
                data['card_expiry_date'] = datetime.strptime(data['card_expiry_date'], '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'تنسيق التاريخ غير صحيح'}), 400
        
        # تحويل التواريخ الجديدة
        date_fields = ['emirates_id_expiry', 'residence_issue_date', 'residence_expiry_date']
        for field in date_fields:
            if data.get(field):
                try:
                    data[field] = datetime.strptime(data[field], '%Y-%m-%d')
                except ValueError:
                    return jsonify({'error': f'تنسيق التاريخ غير صحيح في حقل {field}'}), 400
        
        # إضافة تاريخ الإنشاء
        data['create_date_time'] = datetime.now()
        
        # إدراج الموظف في MongoDB
        result = mongo.db.employees.insert_one(data)
        
        # تسجيل النشاط
        log_activity('إضافة موظف', f'تم إضافة الموظف {data.get("staff_name_ara", "")} - رقم الموظف: {data.get("staff_no", "")}')
        
        # جلب الموظف المضاف مع التفاصيل الكاملة
        employee = mongo.db.employees.find_one({'_id': result.inserted_id})
        emp_dict = serialize_doc(employee)
        emp_dict.update(get_employee_status(employee))
        
        return jsonify(emp_dict), 201
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API لتحديث بيانات موظف
@app.route('/api/employees/<staff_no>', methods=['PUT'])
@require_auth
def update_employee(staff_no):
    try:
        data = request.get_json()
        
        # تحويل تاريخ انتهاء البطاقة إذا كان موجوداً
        if 'card_expiry_date' in data and data['card_expiry_date']:
            try:
                data['card_expiry_date'] = datetime.strptime(data['card_expiry_date'], '%Y-%m-%d')
            except ValueError:
                return jsonify({'error': 'تنسيق التاريخ غير صحيح'}), 400
        elif 'card_expiry_date' in data and not data['card_expiry_date']:
            data['card_expiry_date'] = None
        
        # تحويل التواريخ الجديدة
        date_fields = ['emirates_id_expiry', 'residence_issue_date', 'residence_expiry_date']
        for field in date_fields:
            if field in data and data[field]:
                try:
                    data[field] = datetime.strptime(data[field], '%Y-%m-%d')
                except ValueError:
                    return jsonify({'error': f'تنسيق التاريخ غير صحيح في حقل {field}'}), 400
            elif field in data and not data[field]:
                data[field] = None
        
        # تحديث الموظف - البحث بـ string أولاً، ثم integer للدعم المختلط
        result = mongo.db.employees.update_one(
            {'staff_no': staff_no},
            {'$set': data}
        )
        
        # إذا لم نجد النتيجة بـ string، نجرب بـ integer
        if result.matched_count == 0:
            try:
                staff_no_int = int(staff_no)
                result = mongo.db.employees.update_one(
                    {'staff_no': staff_no_int},
                    {'$set': data}
                )
            except ValueError:
                pass
        
        if result.matched_count == 0:
            return jsonify({'error': 'الموظف غير موجود'}), 404
        
        # تسجيل النشاط
        log_activity('تعديل موظف', f'تم تعديل بيانات الموظف رقم: {staff_no}')
        
        # جلب الموظف المحدث - البحث بـ string أولاً
        employee = mongo.db.employees.find_one({'staff_no': staff_no})
        if not employee:
            try:
                staff_no_int = int(staff_no)
                employee = mongo.db.employees.find_one({'staff_no': staff_no_int})
            except ValueError:
                pass
        
        if not employee:
            return jsonify({'error': 'خطأ في جلب بيانات الموظف المحدث'}), 500
            
        emp_dict = serialize_doc(employee)
        emp_dict.update(get_employee_status(employee))
        
        return jsonify(emp_dict)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API لحذف موظف
@app.route('/api/employees/<staff_no>', methods=['DELETE'])
@require_auth
def delete_employee(staff_no):
    try:
        # جلب بيانات الموظف أولاً للتسجيل
        employee = mongo.db.employees.find_one({'staff_no': staff_no})
        if not employee:
            try:
                staff_no_int = int(staff_no)
                employee = mongo.db.employees.find_one({'staff_no': staff_no_int})
            except ValueError:
                pass
        
        # البحث بـ string أولاً (النوع الصحيح في MongoDB)
        result = mongo.db.employees.delete_one({'staff_no': staff_no})
        
        # إذا لم نجد، نجرب بـ integer للدعم المختلط
        if result.deleted_count == 0:
            try:
                staff_no_int = int(staff_no)
                result = mongo.db.employees.delete_one({'staff_no': staff_no_int})
            except ValueError:
                pass
        
        if result.deleted_count == 0:
            return jsonify({'error': 'الموظف غير موجود'}), 404
        
        # تسجيل النشاط
        employee_name = employee.get('staff_name_ara', employee.get('staff_name', '')) if employee else ''
        log_activity('حذف موظف', f'تم حذف الموظف {employee_name} - رقم الموظف: {staff_no}')
        
        return jsonify({'message': 'تم حذف الموظف بنجاح'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API للحصول على بيانات موظف محدد
@app.route('/api/employees/<staff_no>')
@require_auth
def get_employee(staff_no):
    try:
        # البحث بـ string أولاً (النوع الصحيح في MongoDB)
        employee = mongo.db.employees.find_one({'staff_no': staff_no})
        
        # إذا لم نجد، نجرب بـ integer للدعم المختلط
        if not employee:
            try:
                staff_no_int = int(staff_no)
                employee = mongo.db.employees.find_one({'staff_no': staff_no_int})
            except ValueError:
                pass
        
        if not employee:
            return jsonify({'error': 'الموظف غير موجود'}), 404
        
        emp_dict = serialize_doc(employee)
        emp_dict.update(get_employee_status(employee))
        return jsonify(emp_dict)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API للإحصائيات - بدون مصادقة
@app.route('/api/statistics')
def get_statistics():
    try:
        # التحقق من اتصال MongoDB
        if not mongo:
            return jsonify({
                'error': 'Database connection not available',
                'total_employees': 0,
                'nationality_stats': {},
                'company_stats': {},
                'job_stats': {},
                'passport_missing': 0,
                'cards_missing': 0,
                'cards_expired': 0
            }), 503
        
        total_employees = mongo.db.employees.count_documents({})
    
        # إحصائيات الجنسيات
        nationality_pipeline = [
            {'$group': {'_id': '$nationality_code', 'count': {'$sum': 1}}}
        ]
        nationality_stats = {item['_id']: item['count'] for item in mongo.db.employees.aggregate(nationality_pipeline)}
        
        # إحصائيات الشركات
        company_pipeline = [
            {'$lookup': {
                'from': 'companies',
                'localField': 'company_code',
                'foreignField': 'company_code',
                'as': 'company_info'
            }},
            {'$group': {
                '_id': '$company_code',
                'count': {'$sum': 1},
                'company_name': {'$first': {'$arrayElemAt': ['$company_info.company_name_ara', 0]}}
            }}
        ]
        company_stats = {item['company_name'] or item['_id']: item['count'] 
                        for item in mongo.db.employees.aggregate(company_pipeline)}
        
        # إحصائيات الوظائف
        job_pipeline = [
            {'$lookup': {
                'from': 'jobs',
                'localField': 'job_code',
                'foreignField': 'job_code',
                'as': 'job_info'
            }},
            {'$group': {
                '_id': '$job_code',
                'count': {'$sum': 1},
                'job_name': {'$first': {'$arrayElemAt': ['$job_info.job_ara', 0]}}
            }}
        ]
        job_stats = {item['job_name'] or str(item['_id']): item['count'] 
                    for item in mongo.db.employees.aggregate(job_pipeline)}
        
        # إحصائيات حالة الجوازات والبطاقات
        employees = list(mongo.db.employees.find())
        passport_missing = sum(1 for emp in employees if not emp.get('pass_no'))
        cards_missing = sum(1 for emp in employees if not emp.get('card_no'))
        
        cards_expired = 0
        for emp in employees:
            if emp.get('card_expiry_date'):
                expiry_date = emp['card_expiry_date']
                if isinstance(expiry_date, str):
                    expiry_date = datetime.fromisoformat(expiry_date.replace('Z', '+00:00'))
                # التأكد من أن expiry_date يحتوي على timezone
                if expiry_date.tzinfo is None:
                    expiry_date = expiry_date.replace(tzinfo=timezone.utc)
                if expiry_date < datetime.now(timezone.utc):
                    cards_expired += 1
        
        return jsonify({
            'total_employees': total_employees,
            'nationality_stats': nationality_stats,
            'company_stats': company_stats,
            'job_stats': job_stats,
            'passport_missing': passport_missing,
            'cards_missing': cards_missing,
            'cards_expired': cards_expired
        })
        
    except Exception as e:
        logger.error(f"Error in get_statistics: {e}")
        return jsonify({
            'error': 'خطأ في تحميل الإحصائيات',
            'message': str(e),
            'total_employees': 0,
            'nationality_stats': {},
            'company_stats': {},
            'job_stats': {},
            'passport_missing': 0,
            'cards_missing': 0,
            'cards_expired': 0
        }), 500

# API للحصول على قوائم الفلاتر - بدون مصادقة
@app.route('/api/filters')
def get_filters():
    try:
        # التحقق من اتصال MongoDB
        if not mongo:
            return jsonify({
                'error': 'Database connection not available',
                'nationalities': [],
                'companies': [],
                'jobs': [],
                'departments': []
            }), 503
        
        nationality_codes = mongo.db.employees.distinct('nationality_code')
        
        # تحويل أكواد الجنسيات إلى أسماء كاملة
        nationalities = []
        for code in nationality_codes:
            if code:  # تجاهل القيم الفارغة
                nationalities.append({
                    'code': code,
                    'name_en': get_nationality_name(code, 'en'),
                    'name_ar': get_nationality_name(code, 'ar')
                })
        
        # ترتيب حسب الاسم الإنجليزي
        nationalities.sort(key=lambda x: x['name_en'])
        
        companies = list(mongo.db.companies.find({}, {'company_code': 1, 'company_name_ara': 1, 'company_name_eng': 1, '_id': 0}))
        jobs = list(mongo.db.jobs.find({}, {'job_code': 1, 'job_ara': 1, 'job_eng': 1, '_id': 0}).sort('job_code', 1))
        departments = list(mongo.db.departments.find({}, {'dept_code': 1, 'dept_name_ara': 1, 'dept_name_eng': 1, '_id': 0}).sort('dept_code', 1))
        
        return jsonify({
            'nationalities': nationalities,
            'companies': [{'code': c['company_code'], 'name': c['company_name_ara'], 'name_ara': c['company_name_ara'], 'name_eng': c.get('company_name_eng', c['company_name_ara'])} for c in companies],
            'jobs': [{'code': j['job_code'], 'name': j['job_ara'], 'name_ara': j['job_ara'], 'name_eng': j.get('job_eng', j['job_ara'])} for j in jobs],
            'departments': [{'code': d['dept_code'], 'name': d['dept_name_ara'], 'name_ara': d['dept_name_ara'], 'name_eng': d.get('dept_name_eng', d['dept_name_ara'])} for d in departments]
        })
    except Exception as e:
        logger.error(f"Error in get_filters: {e}")
        return jsonify({
            'error': 'خطأ في تحميل البيانات',
            'message': str(e),
            'nationalities': [],
            'companies': [],
            'jobs': [],
            'departments': []
        }), 500

# APIs للوظائف - قراءة عامة
@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    """جلب الوظائف - لا يحتاج مصادقة"""
    try:
        jobs = list(mongo.db.jobs.find({}, {'_id': 0}).sort('job_code'))
        return jsonify(jobs)
    except Exception as e:
        logger.error(f"Error fetching jobs: {e}")
        return jsonify({'error': 'خطأ في جلب الوظائف'}), 500

@app.route('/api/jobs', methods=['POST'])
@require_auth
def add_job():
    """إضافة وظيفة جديدة - يحتاج مصادقة"""
    try:
        data = request.get_json()
        job_ara = data.get('job_ara', '').strip()
        job_eng = data.get('job_eng', '').strip()
        
        if not job_ara or not job_eng:
            return jsonify({'error': 'اسم الوظيفة بالعربية والإنجليزية مطلوب'}), 400
            
        # التحقق من عدم وجود الوظيفة مسبقاً
        existing_job = mongo.db.jobs.find_one({
            '$or': [
                {'job_ara': job_ara},
                {'job_eng': job_eng}
            ]
        })
        
        if existing_job:
            return jsonify({'error': 'الوظيفة موجودة مسبقاً'}), 400
        
        # إنشاء رقم جديد للوظيفة
        last_job = mongo.db.jobs.find_one({}, sort=[('job_code', -1)])
        new_job_code = int(last_job['job_code']) + 1 if last_job else 1
        
        # إضافة الوظيفة الجديدة
        new_job = {
            'job_code': new_job_code,
            'job_ara': job_ara,
            'job_eng': job_eng,
            'created_at': datetime.now(timezone.utc)
        }
        
        result = mongo.db.jobs.insert_one(new_job)
        
        if result.inserted_id:
            log_activity('ADD_JOB', f'Added new job: {job_ara} / {job_eng}')
            return jsonify({
                'job_code': new_job_code,
                'job_ara': job_ara,
                'job_eng': job_eng
            })
        else:
            return jsonify({'error': 'فشل في إضافة الوظيفة'}), 500
            
    except Exception as e:
        logger.error(f"Error adding new job: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

# API لتعديل وظيفة
@app.route('/api/jobs/<int:job_code>', methods=['PUT'])
@require_auth
def update_job(job_code):
    """تعديل بيانات وظيفة"""
    try:
        data = request.get_json()
        job_ara = data.get('job_ara', '').strip()
        job_eng = data.get('job_eng', '').strip()
        
        if not job_ara or not job_eng:
            return jsonify({'error': 'اسم الوظيفة بالعربية والإنجليزية مطلوب'}), 400
        
        # تحديث الوظيفة
        result = mongo.db.jobs.update_one(
            {'job_code': job_code},
            {'$set': {
                'job_ara': job_ara,
                'job_eng': job_eng,
                'updated_at': datetime.now(timezone.utc)
            }}
        )
        
        if result.matched_count == 0:
            return jsonify({'error': 'الوظيفة غير موجودة'}), 404
        
        log_activity('UPDATE_JOB', f'Updated job: {job_ara} / {job_eng} ({job_code})')
        return jsonify({
            'job_code': job_code,
            'job_ara': job_ara,
            'job_eng': job_eng
        })
            
    except Exception as e:
        logger.error(f"Error updating job: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

# API لحذف وظيفة
@app.route('/api/jobs/<int:job_code>', methods=['DELETE'])
@require_auth
def delete_job(job_code):
    """حذف وظيفة"""
    try:
        # التحقق من وجود موظفين يستخدمون هذه الوظيفة
        employees_count = mongo.db.employees.count_documents({'job_code': job_code})
        if employees_count > 0:
            return jsonify({'error': f'لا يمكن حذف الوظيفة. يوجد {employees_count} موظف مرتبط بهذه الوظيفة'}), 400
        
        # حذف الوظيفة
        result = mongo.db.jobs.delete_one({'job_code': job_code})
        
        if result.deleted_count == 0:
            return jsonify({'error': 'الوظيفة غير موجودة'}), 404
        
        log_activity('DELETE_JOB', f'Deleted job: {job_code}')
        return jsonify({'message': 'تم حذف الوظيفة بنجاح'})
            
    except Exception as e:
        logger.error(f"Error deleting job: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

# API لإدارة الشركات
# APIs للشركات - قراءة عامة
@app.route('/api/companies', methods=['GET'])
def get_companies():
    """جلب الشركات - لا يحتاج مصادقة"""
    try:
        companies = list(mongo.db.companies.find({}, {'_id': 0}).sort('company_name_ara'))
        return jsonify(companies)
    except Exception as e:
        logger.error(f"Error fetching companies: {e}")
        return jsonify({'error': 'خطأ في جلب الشركات'}), 500

@app.route('/api/companies', methods=['POST'])
@require_auth
def add_company():
    """إضافة شركة جديدة - يحتاج مصادقة"""
    try:
        data = request.get_json()
        company_code = data.get('company_code', '').strip().upper()
        company_name_ara = data.get('company_name_ara', '').strip()
        company_name_eng = data.get('company_name_eng', '').strip()
        
        if not company_code or not company_name_ara or not company_name_eng:
            return jsonify({'error': 'جميع بيانات الشركة مطلوبة'}), 400
        
        # التحقق من صحة رمز الشركة
        if not company_code.isalpha():
            return jsonify({'error': 'رمز الشركة يجب أن يحتوي على حروف فقط'}), 400
        
        # التحقق من عدم وجود الشركة مسبقاً
        existing_company = mongo.db.companies.find_one({
            '$or': [
                {'company_code': company_code},
                {'company_name_ara': company_name_ara},
                {'company_name_eng': company_name_eng}
            ]
        })
        
        if existing_company:
            return jsonify({'error': 'الشركة موجودة مسبقاً'}), 400
        
        # إضافة الشركة الجديدة
        new_company = {
            'company_code': company_code,
            'company_name_ara': company_name_ara,
            'company_name_eng': company_name_eng,
            'created_at': datetime.now(timezone.utc)
        }
        
        result = mongo.db.companies.insert_one(new_company)
        
        if result.inserted_id:
            log_activity('ADD_COMPANY', f'Added new company: {company_name_ara} / {company_name_eng} ({company_code})')
            return jsonify({
                'company_code': company_code,
                'company_name_ara': company_name_ara,
                'company_name_eng': company_name_eng
            })
        else:
            return jsonify({'error': 'فشل في إضافة الشركة'}), 500
            
    except Exception as e:
        logger.error(f"Error adding new company: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

# API لتعديل شركة
@app.route('/api/companies/<company_code>', methods=['PUT'])
@require_auth
def update_company(company_code):
    """تعديل بيانات شركة"""
    try:
        data = request.get_json()
        company_name_ara = data.get('company_name_ara', '').strip()
        company_name_eng = data.get('company_name_eng', '').strip()
        
        if not company_name_ara or not company_name_eng:
            return jsonify({'error': 'اسم الشركة بالعربية والإنجليزية مطلوب'}), 400
        
        # تحديث الشركة
        result = mongo.db.companies.update_one(
            {'company_code': company_code},
            {'$set': {
                'company_name_ara': company_name_ara,
                'company_name_eng': company_name_eng,
                'updated_at': datetime.now(timezone.utc)
            }}
        )
        
        if result.matched_count == 0:
            return jsonify({'error': 'الشركة غير موجودة'}), 404
        
        log_activity('UPDATE_COMPANY', f'Updated company: {company_name_ara} / {company_name_eng} ({company_code})')
        return jsonify({
            'company_code': company_code,
            'company_name_ara': company_name_ara,
            'company_name_eng': company_name_eng
        })
            
    except Exception as e:
        logger.error(f"Error updating company: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

# API لحذف شركة
@app.route('/api/companies/<company_code>', methods=['DELETE'])
@require_auth
def delete_company(company_code):
    """حذف شركة"""
    try:
        # التحقق من وجود موظفين يستخدمون هذه الشركة
        employees_count = mongo.db.employees.count_documents({'company_code': company_code})
        if employees_count > 0:
            return jsonify({'error': f'لا يمكن حذف الشركة. يوجد {employees_count} موظف مرتبط بهذه الشركة'}), 400
        
        # حذف الشركة
        result = mongo.db.companies.delete_one({'company_code': company_code})
        
        if result.deleted_count == 0:
            return jsonify({'error': 'الشركة غير موجودة'}), 404
        
        log_activity('DELETE_COMPANY', f'Deleted company: {company_code}')
        return jsonify({'message': 'تم حذف الشركة بنجاح'})
            
    except Exception as e:
        logger.error(f"Error deleting company: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

@app.route('/api/test')
def test_connection():
    """اختبار الاتصال"""
    return jsonify({'status': 'OK', 'message': 'Server is running with MongoDB'})

# API للعرض المختصر للموظفين
@app.route('/api/employees-summary')
def employees_summary():
    """عرض مختصر لجميع الموظفين مع المعلومات الأساسية فقط"""
    query = request.args.get('query', '').strip()
    nationality = request.args.get('nationality', '')
    company = request.args.get('company', '')
    job = request.args.get('job', '')
    department = request.args.get('department', '')
    passport_status = request.args.get('passport_status', '')
    card_status = request.args.get('card_status', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))  # عدد أكبر للعرض المختصر
    
    # بناء الاستعلام (نفس منطق البحث الرئيسي)
    filter_query = {}
    
    if query:
        filter_query['$or'] = [
            {'staff_name': {'$regex': query, '$options': 'i'}},
            {'staff_name_ara': {'$regex': query, '$options': 'i'}},
            {'staff_no': {'$regex': query, '$options': 'i'}}
        ]
    
    if nationality:
        filter_query['nationality_code'] = nationality
    
    if company:
        filter_query['company_code'] = company
    
    if job:
        try:
            filter_query['job_code'] = int(job)
        except (ValueError, TypeError):
            pass  # تجاهل القيم غير الصحيحة
    
    if department:
        filter_query['department_code'] = department
    
    # إضافة فلاتر حالة الجواز والبطاقة (نفس منطق البحث الرئيسي)
    if passport_status == 'missing':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({'$or': [{'pass_no': {'$exists': False}}, {'pass_no': None}, {'pass_no': ''}]})
    elif passport_status == 'available':
        filter_query['pass_no'] = {'$exists': True, '$ne': None, '$ne': ''}
    
    if card_status == 'missing':
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({'$or': [{'card_no': {'$exists': False}}, {'card_no': None}, {'card_no': ''}]})
    elif card_status == 'expired':
        # البطاقات المنتهية الصلاحية (لها تاريخ انتهاء ومنتهية)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'card_no': {'$exists': True, '$ne': None, '$ne': ''},
            'card_expiry_date': {'$exists': True, '$ne': None, '$lt': datetime.now()}
        })
    elif card_status == 'expiring_soon':
        # البطاقات التي ستنتهي خلال 90 يوم
        future_date = datetime.now() + timedelta(days=90)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'card_no': {'$exists': True, '$ne': None, '$ne': ''},
            'card_expiry_date': {'$exists': True, '$ne': None, '$gte': datetime.now(), '$lt': future_date}
        })
    elif card_status == 'valid':
        # البطاقات السارية (لها تاريخ انتهاء ولا تزال سارية)
        future_date = datetime.now() + timedelta(days=90)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'card_no': {'$exists': True, '$ne': None, '$ne': ''},
            'card_expiry_date': {'$exists': True, '$ne': None, '$gte': future_date}
        })
    elif card_status == 'no_expiry':
        # البطاقات بدون تاريخ انتهاء (لها رقم بطاقة ولكن لا يوجد تاريخ انتهاء)
        filter_query['$and'] = filter_query.get('$and', [])
        filter_query['$and'].append({
            'card_no': {'$exists': True, '$ne': None, '$ne': ''},
            '$or': [{'card_expiry_date': {'$exists': False}}, {'card_expiry_date': None}]
        })
    
    # حساب pagination
    skip = (page - 1) * per_page
    
    # جلب البيانات الأساسية فقط
    employees = list(mongo.db.employees.find(
        filter_query,
        {
            'staff_no': 1,
            'staff_name': 1,
            'staff_name_ara': 1,
            'nationality_code': 1,
            'company_code': 1,
            'pass_no': 1,
            'card_no': 1,
            'card_expiry_date': 1
        }
    ).skip(skip).limit(per_page))
    
    total = mongo.db.employees.count_documents(filter_query)
    
    # إعداد النتائج المختصرة
    results = []
    for emp in employees:
        # حالة الجواز
        passport_status = '✅' if emp.get('pass_no') else '❌'
        
        # حالة البطاقة
        if not emp.get('card_no'):
            card_status = '❌'
        elif not emp.get('card_expiry_date'):
            card_status = '⚠️'
        else:
            expiry_date = emp['card_expiry_date']
            if isinstance(expiry_date, str):
                expiry_date = datetime.fromisoformat(expiry_date.replace('Z', '+00:00'))
            
            # التأكد من أن expiry_date يحتوي على timezone
            if expiry_date.tzinfo is None:
                expiry_date = expiry_date.replace(tzinfo=timezone.utc)
            
            if expiry_date < datetime.now(timezone.utc):
                card_status = '🔴'
            elif expiry_date < datetime.now(timezone.utc) + timedelta(days=90):
                card_status = '🟡'
            else:
                card_status = '🟢'
        
        results.append({
            'staff_no': emp.get('staff_no'),
            'name': emp.get('staff_name_ara') or emp.get('staff_name', ''),
            'nationality': emp.get('nationality_code', ''),
            'company': emp.get('company_code', ''),
            'passport': passport_status,
            'card': card_status
        })
    
    pages = (total + per_page - 1) // per_page
    
    return jsonify({
        'employees': results,
        'total': total,
        'pages': pages,
        'current_page': page,
        'has_next': page < pages,
        'has_prev': page > 1
    })

# API تصدير النتائج المفلترة
@app.route('/api/get-detailed-results', methods=['POST'])
@require_auth
def get_detailed_results():
    """جلب النتائج التفصيلية للتصدير"""
    try:
        data = request.get_json()
        
        # التحقق من وجود البيانات
        if not data:
            logger.warning("No data received in get_detailed_results")
            return jsonify({'error': 'لا توجد بيانات مرسلة'}), 400
        
        # معايير البحث
        query = data.get('query', '').strip()
        nationality = data.get('nationality', '')
        company = data.get('company', '')
        job = data.get('job', '')
        passport_status = data.get('passport_status', '')
        card_status = data.get('card_status', '')
        emirates_id_status = data.get('emirates_id_status', '')
        residence_status = data.get('residence_status', '')
        
        logger.info(f"Detailed search request: query='{query}', nationality='{nationality}', company='{company}', job='{job}', passport_status='{passport_status}', card_status='{card_status}', emirates_id_status='{emirates_id_status}', residence_status='{residence_status}'")
        
        # بناء استعلام البحث (نفس المنطق من API البحث الأساسي)
        filter_query = {}
        
        if query:
            regex_pattern = '.*' + re.escape(query) + '.*'
            filter_query['$or'] = [
                {'staff_name': {'$regex': regex_pattern, '$options': 'i'}},
                {'staff_name_ara': {'$regex': regex_pattern, '$options': 'i'}},
                {'staff_no': {'$regex': regex_pattern, '$options': 'i'}},
                {'pass_no': {'$regex': regex_pattern, '$options': 'i'}},
                {'card_no': {'$regex': regex_pattern, '$options': 'i'}},
                {'emirates_id': {'$regex': regex_pattern, '$options': 'i'}},
                {'residence_no': {'$regex': regex_pattern, '$options': 'i'}}
            ]
        
        if nationality:
            filter_query['nationality_code'] = nationality
        
        if company:
            filter_query['company_code'] = company
            
        if job:
            filter_query['job_code'] = job
        
        # فلترة حالة الجواز
        if passport_status == 'available':
            filter_query['pass_no'] = {'$exists': True, '$ne': None, '$ne': ''}
        elif passport_status == 'missing':
            if '$or' in filter_query:
                # إذا كان هناك $or مسبقاً (من البحث النصي)، نحوله إلى $and
                existing_or = filter_query.pop('$or')
                filter_query['$and'] = filter_query.get('$and', [])
                filter_query['$and'].append({'$or': existing_or})
                filter_query['$and'].append({
                    '$or': [
                        {'pass_no': {'$exists': False}},
                        {'pass_no': None},
                        {'pass_no': ''}
                    ]
                })
            else:
                filter_query['$or'] = [
                    {'pass_no': {'$exists': False}},
                    {'pass_no': None},
                    {'pass_no': ''}
                ]
        
        # فلترة حالة البطاقة
        if card_status == 'valid':
            current_date = datetime.now(timezone.utc)
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'card_no': {'$exists': True, '$ne': None, '$ne': ''},
                'card_expiry_date': {'$gte': current_date + timedelta(days=90)}
            })
        elif card_status == 'expiring':
            current_date = datetime.now(timezone.utc)
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'card_no': {'$exists': True, '$ne': None, '$ne': ''},
                'card_expiry_date': {
                    '$gte': current_date,
                    '$lt': current_date + timedelta(days=90)
                }
            })
        elif card_status == 'expired':
            current_date = datetime.now(timezone.utc)
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'card_no': {'$exists': True, '$ne': None, '$ne': ''},
                'card_expiry_date': {'$lt': current_date}
            })
        elif card_status == 'missing':
            if '$or' in filter_query:
                # إذا كان هناك $or مسبقاً، نحوله إلى $and
                existing_or = filter_query.pop('$or')
                filter_query['$and'] = filter_query.get('$and', [])
                filter_query['$and'].append({'$or': existing_or})
                filter_query['$and'].append({
                    '$or': [
                        {'card_no': {'$exists': False}},
                        {'card_no': None},
                        {'card_no': ''}
                    ]
                })
            else:
                filter_query['$or'] = [
                    {'card_no': {'$exists': False}},
                    {'card_no': None},
                    {'card_no': ''}
                ]
        elif card_status == 'no_expiry':
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'card_no': {'$exists': True, '$ne': None, '$ne': ''},
                '$or': [{'card_expiry_date': {'$exists': False}}, {'card_expiry_date': None}]
            })
        
        # إضافة فلاتر حالة الهوية الإماراتية
        if emirates_id_status == 'missing':
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({'$or': [{'emirates_id': {'$exists': False}}, {'emirates_id': None}, {'emirates_id': ''}]})
        elif emirates_id_status == 'expired':
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'emirates_id': {'$exists': True, '$ne': None, '$ne': ''},
                'emirates_id_expiry': {'$exists': True, '$ne': None, '$lt': datetime.now()}
            })
        elif emirates_id_status == 'expiring_soon':
            future_date = datetime.now() + timedelta(days=90)
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'emirates_id': {'$exists': True, '$ne': None, '$ne': ''},
                'emirates_id_expiry': {'$exists': True, '$ne': None, '$gte': datetime.now(), '$lt': future_date}
            })
        elif emirates_id_status == 'valid':
            future_date = datetime.now() + timedelta(days=90)
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'emirates_id': {'$exists': True, '$ne': None, '$ne': ''},
                'emirates_id_expiry': {'$exists': True, '$ne': None, '$gte': future_date}
            })
        elif emirates_id_status == 'no_expiry':
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'emirates_id': {'$exists': True, '$ne': None, '$ne': ''},
                '$or': [{'emirates_id_expiry': {'$exists': False}}, {'emirates_id_expiry': None}]
            })
        
        # إضافة فلاتر حالة الإقامة
        if residence_status == 'missing':
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({'$or': [{'residence_no': {'$exists': False}}, {'residence_no': None}, {'residence_no': ''}]})
        elif residence_status == 'expired':
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'residence_no': {'$exists': True, '$ne': None, '$ne': ''},
                'residence_expiry_date': {'$exists': True, '$ne': None, '$lt': datetime.now()}
            })
        elif residence_status == 'expiring_soon':
            future_date = datetime.now() + timedelta(days=90)
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'residence_no': {'$exists': True, '$ne': None, '$ne': ''},
                'residence_expiry_date': {'$exists': True, '$ne': None, '$gte': datetime.now(), '$lt': future_date}
            })
        elif residence_status == 'valid':
            future_date = datetime.now() + timedelta(days=90)
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'residence_no': {'$exists': True, '$ne': None, '$ne': ''},
                'residence_expiry_date': {'$exists': True, '$ne': None, '$gte': future_date}
            })
        elif residence_status == 'no_expiry':
            filter_query['$and'] = filter_query.get('$and', [])
            filter_query['$and'].append({
                'residence_no': {'$exists': True, '$ne': None, '$ne': ''},
                '$or': [{'residence_expiry_date': {'$exists': False}}, {'residence_expiry_date': None}]
            })
        
        # جلب جميع البيانات التفصيلية
        try:
            employees = list(mongo.db.employees.find(filter_query))
            logger.info(f"Found {len(employees)} employees matching filter: {filter_query}")
        except Exception as db_error:
            logger.error(f"Database query error: {str(db_error)}")
            return jsonify({'error': f'خطأ في قاعدة البيانات: {str(db_error)}'}), 500
        
        # معالجة البيانات وإضافة نصوص الحالة
        detailed_results = []
        for emp in employees:
            # تحويل ObjectId إلى string
            if '_id' in emp:
                emp['_id'] = str(emp['_id'])
            
            # حالة الجواز
            if emp.get('pass_no'):
                passport_text = 'متوفر'
            else:
                passport_text = 'غير متوفر'
            
            # حالة البطاقة  
            if not emp.get('card_no'):
                card_text = 'مفقودة'
            elif not emp.get('card_expiry_date'):
                card_text = 'بدون تاريخ انتهاء'
            else:
                expiry_date = emp['card_expiry_date']
                if isinstance(expiry_date, str):
                    expiry_date = datetime.fromisoformat(expiry_date.replace('Z', '+00:00'))
                
                if expiry_date.tzinfo is None:
                    expiry_date = expiry_date.replace(tzinfo=timezone.utc)
                
                current_date = datetime.now(timezone.utc)
                if expiry_date < current_date:
                    card_text = 'منتهية'
                elif expiry_date < current_date + timedelta(days=90):
                    card_text = 'تنتهي قريباً'
                else:
                    card_text = 'سارية'
            
            # إضافة البيانات مع النصوص
            emp_data = emp.copy()
            # تحويل ObjectId إلى string
            if '_id' in emp_data:
                emp_data['_id'] = str(emp_data['_id'])
            emp_data['passport_text'] = passport_text
            emp_data['card_text'] = card_text
            
            # تنسيق التواريخ
            if emp_data.get('card_expiry_date'):
                if isinstance(emp_data['card_expiry_date'], datetime):
                    emp_data['card_expiry_date'] = emp_data['card_expiry_date'].strftime('%Y-%m-%d')
            
            if emp_data.get('emirates_id_expiry'):
                if isinstance(emp_data['emirates_id_expiry'], datetime):
                    emp_data['emirates_id_expiry'] = emp_data['emirates_id_expiry'].strftime('%Y-%m-%d')
                    
            if emp_data.get('residence_issue_date'):
                if isinstance(emp_data['residence_issue_date'], datetime):
                    emp_data['residence_issue_date'] = emp_data['residence_issue_date'].strftime('%Y-%m-%d')
                    
            if emp_data.get('residence_expiry_date'):
                if isinstance(emp_data['residence_expiry_date'], datetime):
                    emp_data['residence_expiry_date'] = emp_data['residence_expiry_date'].strftime('%Y-%m-%d')
            
            detailed_results.append(emp_data)
        
        return jsonify({
            'employees': detailed_results,
            'total': len(detailed_results)
        })
        
    except Exception as e:
        logger.error(f"Error in get_detailed_results: {str(e)}")
        logger.error(f"Filter query was: {filter_query if 'filter_query' in locals() else 'Not defined'}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

@app.route('/api/export-filtered-results', methods=['POST'])
@require_auth
def export_filtered_results():
    """تصدير النتائج المفلترة كملف Excel"""
    try:
        data = request.get_json()
        employees = data.get('employees', [])
        filters = data.get('filters', {})
        total = data.get('total', 0)
        
        if not employees:
            return jsonify({'error': 'لا توجد بيانات للتصدير'}), 400
        
        # إعداد البيانات للتصدير
        report_data = []
        for emp in employees:
            # جلب معلومات الشركة والوظيفة
            company_info = mongo.db.companies.find_one({'company_code': emp.get('company_code')})
            job_info = mongo.db.jobs.find_one({'job_code': emp.get('job_code')})
            
            report_data.append({
                'رقم الموظف': emp.get('staff_no', ''),
                'الاسم بالعربية': emp.get('staff_name_ara', ''),
                'الاسم بالإنجليزية': emp.get('staff_name', ''),
                'الوظيفة': job_info.get('job_ara', '') if job_info else '',
                'الشركة': company_info.get('company_name_ara', '') if company_info else '',
                'الجنسية': emp.get('nationality_code', ''),
                'رقم الجواز': emp.get('pass_no', 'غير متوفر'),
                'حالة الجواز': emp.get('passport_text', ''),
                'رقم البطاقة': emp.get('card_no', 'غير متوفرة'),
                'حالة البطاقة': emp.get('card_text', ''),
                'تاريخ انتهاء البطاقة': emp.get('card_expiry_date', 'غير محدد'),
                'رقم الهوية الإماراتية': emp.get('emirates_id', 'غير متوفر'),
                'تاريخ انتهاء الهوية': emp.get('emirates_id_expiry', 'غير محدد'),
                'رقم الإقامة': emp.get('residence_no', 'غير متوفر'),
                'تاريخ إصدار الإقامة': emp.get('residence_issue_date', 'غير محدد'),
                'تاريخ انتهاء الإقامة': emp.get('residence_expiry_date', 'غير محدد'),
                'تاريخ الإنشاء': emp.get('create_date_time', '')
            })
        
        # إنشاء DataFrame
        df = pd.DataFrame(report_data)
        
        # إنشاء ملف Excel في الذاكرة
        output = io.BytesIO()
        try:
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # كتابة البيانات
                df.to_excel(writer, sheet_name='بيانات الموظفين', index=False, startrow=3)
                
                # تنسيق ورقة بيانات الموظفين
                worksheet = writer.sheets['بيانات الموظفين']
                
                # استيراد التنسيقات من openpyxl
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                
                # إضافة عنوان رئيسي للتقرير
                worksheet.merge_cells('A1:Q2')  # تعديل النطاق ليشمل الأعمدة الجديدة
                title_cell = worksheet['A1']
                title_cell.value = f"تقرير بيانات الموظفين - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                title_cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                title_cell.border = Border(
                    left=Side(style='thick', color='000000'),
                    right=Side(style='thick', color='000000'),
                    top=Side(style='thick', color='000000'),
                    bottom=Side(style='thick', color='000000')
                )
                
                # تنسيق العنوان (الصف الرابع)
                header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
                header_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                # تطبيق تنسيق العنوان
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=4, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = header_border
                    cell.alignment = header_alignment
                
                # تنسيق بيانات الجدول
                data_font = Font(name='Calibri', size=11)
                data_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                data_alignment = Alignment(horizontal='center', vertical='center')
                
                # تطبيق تنسيق البيانات مع تلوين متناوب للصفوف
                for row in range(5, len(df) + 5):
                    # تلوين متناوب للصفوف
                    if row % 2 == 0:
                        row_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                    else:
                        row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = data_font
                        cell.border = data_border
                        cell.alignment = data_alignment
                        
                        # تلوين خاص لحالات الجواز والبطاقة
                        col_name = df.columns[col-1]
                        cell_value = str(cell.value) if cell.value else ''
                        
                        if col_name == 'حالة الجواز':
                            if 'غير متوفر' in cell_value or 'مفقود' in cell_value:
                                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # أحمر فاتح
                                cell.font = Font(name='Calibri', size=11, color='CC0000', bold=True)
                            elif 'متوفر' in cell_value or 'موجود' in cell_value:
                                cell.fill = PatternFill(start_color='E6F7E6', end_color='E6F7E6', fill_type='solid')  # أخضر فاتح
                                cell.font = Font(name='Calibri', size=11, color='008000', bold=True)
                            else:
                                cell.fill = row_fill
                        elif col_name == 'حالة البطاقة':
                            if 'منتهية' in cell_value or 'مفقودة' in cell_value:
                                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # أحمر فاتح
                                cell.font = Font(name='Calibri', size=11, color='CC0000', bold=True)
                            elif 'تنتهي قريباً' in cell_value:
                                cell.fill = PatternFill(start_color='FFF2E6', end_color='FFF2E6', fill_type='solid')  # برتقالي فاتح
                                cell.font = Font(name='Calibri', size=11, color='FF8000', bold=True)
                            elif 'سارية' in cell_value:
                                cell.fill = PatternFill(start_color='E6F7E6', end_color='E6F7E6', fill_type='solid')  # أخضر فاتح
                                cell.font = Font(name='Calibri', size=11, color='008000', bold=True)
                            else:
                                cell.fill = row_fill
                        else:
                            cell.fill = row_fill
                
                # تعديل عرض الأعمدة تلقائياً
                for col in range(1, len(df.columns) + 1):
                    column_letter = get_column_letter(col)
                    max_length = max(len(str(df.iloc[row, col-1])) for row in range(len(df)))
                    header_length = len(df.columns[col-1])
                    adjusted_width = max(max_length, header_length) + 2
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 30)
                
                # إضافة معلومات الفلتر
                filter_info = []
                if filters.get('query'):
                    filter_info.append(f"البحث: {filters['query']}")
                if filters.get('nationality'):
                    filter_info.append(f"الجنسية: {filters['nationality']}")
                if filters.get('company'):
                    filter_info.append(f"الشركة: {filters['company']}")
                if filters.get('passport_status'):
                    filter_info.append(f"حالة الجواز: {filters['passport_status']}")
                if filters.get('card_status'):
                    filter_info.append(f"حالة البطاقة: {filters['card_status']}")
                
                # إضافة ورقة معلومات التقرير
                summary_data = [
                    ['إجمالي النتائج', total],
                    ['تاريخ التقرير', datetime.now().strftime('%Y-%m-%d %H:%M')],
                    ['الفلاتر المطبقة', ' | '.join(filter_info) if filter_info else 'لا توجد فلاتر']
                ]
                
                summary_df = pd.DataFrame(summary_data, columns=['البيان', 'القيمة'])
                summary_df.to_excel(writer, sheet_name='معلومات التقرير', index=False)
                
                # تنسيق ورقة معلومات التقرير
                summary_worksheet = writer.sheets['معلومات التقرير']
                
                # تنسيق عنوان ورقة معلومات التقرير
                summary_header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
                summary_header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                
                for col in range(1, 3):
                    cell = summary_worksheet.cell(row=1, column=col)
                    cell.font = summary_header_font
                    cell.fill = summary_header_fill
                    cell.border = header_border
                    cell.alignment = header_alignment
                
                # تنسيق بيانات ورقة معلومات التقرير
                summary_data_font = Font(name='Calibri', size=12)
                for row in range(2, len(summary_data) + 2):
                    for col in range(1, 3):
                        cell = summary_worksheet.cell(row=row, column=col)
                        cell.font = summary_data_font
                        cell.border = data_border
                        cell.alignment = data_alignment
                        if col == 1:  # عمود البيان
                            cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
                        else:  # عمود القيمة
                            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                # تعديل عرض أعمدة ورقة معلومات التقرير
                summary_worksheet.column_dimensions['A'].width = 20
                summary_worksheet.column_dimensions['B'].width = 30
                
        except Exception as e:
            return jsonify({'error': f'خطأ في إنشاء ملف Excel: {str(e)}'}), 500
            
        output.seek(0)
        
        # إرسال الملف
        filename = f"تقرير_موظفين_مفلتر_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API لجلب سجل التدقيق
@app.route('/api/audit-logs')
def get_audit_logs():
    """جلب سجل التدقيق مع الترقيم"""
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        action_filter = request.args.get('action', '')
        
        # بناء الاستعلام
        filter_query = {}
        if action_filter:
            filter_query['action'] = {'$regex': action_filter, '$options': 'i'}
        
        # حساب إجمالي السجلات
        total = mongo.db.audit_logs.count_documents(filter_query)
        
        # جلب البيانات مع الترقيم
        logs = list(mongo.db.audit_logs.find(filter_query)
                   .sort('timestamp', -1)
                   .skip((page - 1) * per_page)
                   .limit(per_page))
        
        # تنسيق البيانات
        formatted_logs = []
        for log in logs:
            formatted_logs.append({
                'timestamp': log.get('timestamp').strftime('%Y-%m-%d %H:%M:%S') if log.get('timestamp') else '',
                'action': log.get('action', ''),
                'details': log.get('details', ''),
                'user_id': log.get('user_id', 'مجهول'),
                'ip_address': log.get('ip_address', '')
            })
        
        return jsonify({
            'logs': formatted_logs,
            'total': total,
            'page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# API لحذف سجلات التدقيق القديمة
@app.route('/api/audit-logs/cleanup', methods=['POST'])
@require_auth
def cleanup_audit_logs():
    """حذف سجلات التدقيق الأقدم من 90 يوم"""
    try:
        cutoff_date = datetime.now() - timedelta(days=90)
        result = mongo.db.audit_logs.delete_many({'timestamp': {'$lt': cutoff_date}})
        
        log_activity('تنظيف سجل التدقيق', f'تم حذف {result.deleted_count} سجل قديم')
        
        return jsonify({
            'success': True,
            'deleted_count': result.deleted_count,
            'message': f'تم حذف {result.deleted_count} سجل قديم'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# APIs للأقسام
@app.route('/api/departments')
def get_departments():
    """جلب جميع الأقسام - لا يحتاج مصادقة"""
    try:
        departments = list(mongo.db.departments.find({}, {'_id': 0}).sort('department_name_ara'))
        return jsonify(departments)
    except Exception as e:
        logger.error(f"Error fetching departments: {e}")
        return jsonify({'error': 'خطأ في جلب الأقسام'}), 500

# APIs عامة للبيانات الأساسية (بدون مصادقة)
@app.route('/api/public/jobs')
def get_public_jobs():
    """جلب الوظائف - نسخة عامة بدون مصادقة"""
    try:
        jobs = list(mongo.db.jobs.find({}, {'_id': 0}).sort('job_code'))
        return jsonify(jobs)
    except Exception as e:
        logger.error(f"Error fetching public jobs: {e}")
        return jsonify({'error': 'خطأ في جلب الوظائف'}), 500

@app.route('/api/public/companies')
def get_public_companies():
    """جلب الشركات - نسخة عامة بدون مصادقة"""
    try:
        companies = list(mongo.db.companies.find({}, {'_id': 0}).sort('company_name_ara'))
        return jsonify(companies)
    except Exception as e:
        logger.error(f"Error fetching public companies: {e}")
        return jsonify({'error': 'خطأ في جلب الشركات'}), 500

@app.route('/api/public/departments')
def get_public_departments():
    """جلب الأقسام - نسخة عامة بدون مصادقة"""
    try:
        departments = list(mongo.db.departments.find({}, {'_id': 0}).sort('department_name_ara'))
        return jsonify(departments)
    except Exception as e:
        logger.error(f"Error fetching public departments: {e}")
        return jsonify({'error': 'خطأ في جلب الأقسام'}), 500

@app.route('/api/departments', methods=['POST'])
@require_auth
def add_department():
    """إضافة قسم جديد"""
    try:
        data = request.get_json()
        
        # التحقق من البيانات المطلوبة
        if not data.get('department_code') or not data.get('department_name_ara') or not data.get('department_name_eng'):
            return jsonify({'error': 'جميع حقول القسم مطلوبة'}), 400
        
        # التحقق من عدم وجود قسم بنفس الكود
        existing = mongo.db.departments.find_one({'department_code': data['department_code']})
        if existing:
            return jsonify({'error': 'كود القسم موجود مسبقاً'}), 400
        
        # إدراج القسم الجديد
        mongo.db.departments.insert_one(data)
        
        log_activity('إضافة قسم', f'تم إضافة القسم {data.get("department_name_ara", "")} - كود: {data.get("department_code", "")}')
        
        return jsonify(data), 201
        
    except Exception as e:
        logger.error(f"Error adding department: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

@app.route('/api/departments/<department_code>', methods=['PUT'])
@require_auth
def update_department(department_code):
    """تحديث قسم"""
    try:
        data = request.get_json()
        
        # تحديث القسم
        result = mongo.db.departments.update_one(
            {'department_code': department_code},
            {'$set': {
                'department_name_ara': data.get('department_name_ara'),
                'department_name_eng': data.get('department_name_eng')
            }}
        )
        
        if result.matched_count == 0:
            return jsonify({'error': 'القسم غير موجود'}), 404
        
        log_activity('تحديث قسم', f'تم تحديث القسم: {data.get("department_name_ara", "")} ({department_code})')
        return jsonify(data)
            
    except Exception as e:
        logger.error(f"Error updating department: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

@app.route('/api/departments/<department_code>', methods=['DELETE'])
@require_auth
def delete_department(department_code):
    """حذف قسم"""
    try:
        # التحقق من وجود موظفين يستخدمون هذا القسم
        employees_count = mongo.db.employees.count_documents({'department_code': department_code})
        if employees_count > 0:
            return jsonify({'error': f'لا يمكن حذف القسم. يوجد {employees_count} موظف مرتبط بهذا القسم'}), 400
        
        # حذف القسم
        result = mongo.db.departments.delete_one({'department_code': department_code})
        
        if result.deleted_count == 0:
            return jsonify({'error': 'القسم غير موجود'}), 404
        
        log_activity('حذف قسم', f'تم حذف القسم: {department_code}')
        return jsonify({'message': 'تم حذف القسم بنجاح'})
            
    except Exception as e:
        logger.error(f"Error deleting department: {e}")
        return jsonify({'error': f'خطأ في الخادم: {str(e)}'}), 500

def load_initial_data():
    """تحميل البيانات الأولية في MongoDB"""
    
    # التحقق من وجود البيانات
    if mongo.db.companies.count_documents({}) > 0:
        print("البيانات محملة مسبقاً!")
        return
    
    # بيانات الشركات
    companies_data = [
        {"company_code": "BRG", "company_name_eng": "Middle East Bridge Trading", "company_name_ara": "میدل ایست بریدج للتجارة العامة (ش.ذ.م.م)"},
        {"company_code": "HON", "company_name_eng": "Honda Resources", "company_name_ara": "موارد هوندا"},
        {"company_code": "LIV", "company_name_eng": "Liverage Trading", "company_name_ara": "ليفردج للتجارة"},
        {"company_code": "MNT", "company_name_eng": "Mınt Art Galery", "company_name_ara": "مينت آرت جالاري"},
        {"company_code": "SQF", "company_name_eng": "SQFT General Store", "company_name_ara": "اس كيو اف تي للمخازن العامة"},
        {"company_code": "TAM", "company_name_eng": "Tamayoz", "company_name_ara": "تميز"},
        {"company_code": "UNI", "company_name_eng": "UNI FOOD GENERAL TRADING LLC", "company_name_ara": "يونيفود للتجارة"}
    ]
    
    mongo.db.companies.insert_many(companies_data)
    
    # بيانات الوظائف
    jobs_data = [
        {"job_code": 1, "job_eng": "Accountant", "job_ara": "محاسب"},
        {"job_code": 2, "job_eng": "Archive Clerk", "job_ara": "كاتب الأرشیف"},
        {"job_code": 3, "job_eng": "Commercial Sales Representative", "job_ara": "ممثل مبيعات تجاري"},
        {"job_code": 4, "job_eng": "Computer Engineer", "job_ara": "مھندس كومبیوتر"},
        {"job_code": 5, "job_eng": "Filing Clerk", "job_ara": "كاتب ملفات"},
        {"job_code": 6, "job_eng": "Marketing Manager", "job_ara": "مدير التسويق"},
        {"job_code": 7, "job_eng": "Messenger", "job_ara": "مراسل"},
        {"job_code": 8, "job_eng": "Operations Manager", "job_ara": "مدير عمليات"},
        {"job_code": 9, "job_eng": "Sales Manager", "job_ara": "مدیر المبیعات"},
        {"job_code": 10, "job_eng": "Shop Assistant", "job_ara": "عامل مساعد بمتجر"},
        {"job_code": 11, "job_eng": "Stall and Market Salesperson", "job_ara": "مندوب مبيعات الأكشاك والسوق"},
        {"job_code": 12, "job_eng": "Stevedore", "job_ara": "محمل سفن"},
        {"job_code": 13, "job_eng": "Legal Consultant", "job_ara": "استشاري قانوني"},
        {"job_code": 14, "job_eng": "Finance Director", "job_ara": "مدیر المالیة"},
        {"job_code": 15, "job_eng": "Administration Manager", "job_ara": "مدير ادارة"},
        {"job_code": 16, "job_eng": "Loading and unloading worker", "job_ara": "عامل الشحن والتفريغ"},
        {"job_code": 17, "job_eng": "Marketing Specialist", "job_ara": "أخصائي تسويق"},
        {"job_code": 18, "job_eng": "Storekeeper", "job_ara": "أمين مخزن"},
        {"job_code": 19, "job_eng": "General Manager", "job_ara": "مدير عام"}
    ]
    
    mongo.db.jobs.insert_many(jobs_data)
    
    # بيانات الأقسام
    departments_data = [
        {"department_code": "HR", "department_name_eng": "Human Resources", "department_name_ara": "الموارد البشرية"},
        {"department_code": "FIN", "department_name_eng": "Finance", "department_name_ara": "المالية"},
        {"department_code": "IT", "department_name_eng": "Information Technology", "department_name_ara": "تقنية المعلومات"},
        {"department_code": "MKT", "department_name_eng": "Marketing", "department_name_ara": "التسويق"},
        {"department_code": "SLS", "department_name_eng": "Sales", "department_name_ara": "المبيعات"},
        {"department_code": "OPS", "department_name_eng": "Operations", "department_name_ara": "العمليات"},
        {"department_code": "LOG", "department_name_eng": "Logistics", "department_name_ara": "الخدمات اللوجستية"},
        {"department_code": "ADM", "department_name_eng": "Administration", "department_name_ara": "الإدارة"},
        {"department_code": "LEG", "department_name_eng": "Legal", "department_name_ara": "الشؤون القانونية"},
        {"department_code": "WAR", "department_name_eng": "Warehouse", "department_name_ara": "المخازن"}
    ]
    
    # حذف البيانات الموجودة وإدراج البيانات الجديدة
    mongo.db.departments.delete_many({})
    mongo.db.departments.insert_many(departments_data)
    
    print("تم تحميل بيانات الشركات والوظائف والأقسام بنجاح في MongoDB!")
    
    # إنشاء المستخدم الإداري
    init_admin_user()

if __name__ == '__main__':
    # تحميل البيانات الأولية
    load_initial_data()
    
    port = int(os.getenv('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
